from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from .models import Medico, Paciente, Cita, Receta, Especialidad, ConsultaReporte
from .forms import MedicoForm, PacienteForm, CitaForm, RecetaForm, EspecialidadForm
from django.contrib import messages
from django.db import connection
import re

# Importaciones para exportación
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io
from datetime import datetime

@login_required
def dashboard(request):
    context = {
        'medicos': Medico.objects.all(),
        'pacientes': Paciente.objects.all(),
        'citas': Cita.objects.all(),
        'recetas': Receta.objects.all(),
        'especialidades': Especialidad.objects.all(),
    }
    return render(request, 'dashboard.html', context)

# CRUD para Especialidad
@login_required
def especialidad_create(request):
    if request.method == 'POST':
        form = EspecialidadForm(request.POST)
        if form.is_valid():
            form.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Especialidad creada correctamente'})
            messages.success(request, 'Especialidad creada correctamente')
            return redirect('dashboard')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': form.errors})
    else:
        form = EspecialidadForm()
    return render(request, 'gestion/form_modal.html', {'form': form, 'title': 'Crear Especialidad'})

@login_required
def especialidad_update(request, pk):
    especialidad = get_object_or_404(Especialidad, pk=pk)
    if request.method == 'POST':
        form = EspecialidadForm(request.POST, instance=especialidad)
        if form.is_valid():
            form.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Especialidad actualizada correctamente'})
            messages.success(request, 'Especialidad actualizada correctamente')
            return redirect('dashboard')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': form.errors})
    else:
        form = EspecialidadForm(instance=especialidad)
    return render(request, 'gestion/form_modal.html', {'form': form, 'title': 'Editar Especialidad'})

@login_required
def especialidad_delete(request, pk):
    especialidad = get_object_or_404(Especialidad, pk=pk)
    if request.method == 'POST':
        especialidad.delete()
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': 'Especialidad eliminada correctamente'})
        messages.success(request, 'Especialidad eliminada correctamente')
        return redirect('dashboard')
    return render(request, 'gestion/confirm_delete.html', {'object': especialidad, 'title': 'Confirmar eliminación de Especialidad'})

# CRUD para Medico
@login_required
def medico_create(request):
    if request.method == 'POST':
        form = MedicoForm(request.POST)
        if form.is_valid():
            form.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Médico creado correctamente'})
            messages.success(request, 'Médico creado correctamente')
            return redirect('dashboard')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': form.errors})
    else:
        form = MedicoForm()
    return render(request, 'gestion/form_modal.html', {'form': form, 'title': 'Crear Médico'})

@login_required
def medico_update(request, pk):
    medico = get_object_or_404(Medico, pk=pk)
    if request.method == 'POST':
        form = MedicoForm(request.POST, instance=medico)
        if form.is_valid():
            form.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Médico actualizado correctamente'})
            messages.success(request, 'Médico actualizado correctamente')
            return redirect('dashboard')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': form.errors})
    else:
        form = MedicoForm(instance=medico)
    return render(request, 'gestion/form_modal.html', {'form': form, 'title': 'Editar Médico'})

@login_required
def medico_delete(request, pk):
    medico = get_object_or_404(Medico, pk=pk)
    if request.method == 'POST':
        medico.delete()
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': 'Médico eliminado correctamente'})
        messages.success(request, 'Médico eliminado correctamente')
        return redirect('dashboard')
    return render(request, 'gestion/confirm_delete.html', {'object': medico, 'title': 'Confirmar eliminación de Médico'})

# CRUD para Paciente
@login_required
def paciente_create(request):
    if request.method == 'POST':
        form = PacienteForm(request.POST)
        if form.is_valid():
            form.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Paciente creado correctamente'})
            messages.success(request, 'Paciente creado correctamente')
            return redirect('dashboard')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': form.errors})
    else:
        form = PacienteForm()
    return render(request, 'gestion/form_modal.html', {'form': form, 'title': 'Crear Paciente'})

@login_required
def paciente_update(request, pk):
    paciente = get_object_or_404(Paciente, pk=pk)
    if request.method == 'POST':
        form = PacienteForm(request.POST, instance=paciente)
        if form.is_valid():
            form.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Paciente actualizado correctamente'})
            messages.success(request, 'Paciente actualizado correctamente')
            return redirect('dashboard')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': form.errors})
    else:
        form = PacienteForm(instance=paciente)
    return render(request, 'gestion/form_modal.html', {'form': form, 'title': 'Editar Paciente'})

@login_required
def paciente_delete(request, pk):
    paciente = get_object_or_404(Paciente, pk=pk)
    if request.method == 'POST':
        paciente.delete()
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': 'Paciente eliminado correctamente'})
        messages.success(request, 'Paciente eliminado correctamente')
        return redirect('dashboard')
    return render(request, 'gestion/confirm_delete.html', {'object': paciente, 'title': 'Confirmar eliminación de Paciente'})

# CRUD para Cita
@login_required
def cita_create(request):
    if request.method == 'POST':
        form = CitaForm(request.POST)
        if form.is_valid():
            form.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Cita creada correctamente'})
            messages.success(request, 'Cita creada correctamente')
            return redirect('dashboard')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': form.errors})
    else:
        form = CitaForm()
    return render(request, 'gestion/form_modal.html', {'form': form, 'title': 'Crear Cita'})

@login_required
def cita_update(request, pk):
    cita = get_object_or_404(Cita, pk=pk)
    if request.method == 'POST':
        form = CitaForm(request.POST, instance=cita)
        if form.is_valid():
            form.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Cita actualizada correctamente'})
            messages.success(request, 'Cita actualizada correctamente')
            return redirect('dashboard')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': form.errors})
    else:
        form = CitaForm(instance=cita)
    return render(request, 'gestion/form_modal.html', {'form': form, 'title': 'Editar Cita'})

@login_required
def cita_delete(request, pk):
    cita = get_object_or_404(Cita, pk=pk)
    if request.method == 'POST':
        cita.delete()
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': 'Cita eliminada correctamente'})
        messages.success(request, 'Cita eliminada correctamente')
        return redirect('dashboard')
    return render(request, 'gestion/confirm_delete.html', {'object': cita, 'title': 'Confirmar eliminación de Cita'})

# CRUD para Receta
@login_required
def receta_create(request):
    if request.method == 'POST':
        form = RecetaForm(request.POST)
        if form.is_valid():
            form.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Receta creada correctamente'})
            messages.success(request, 'Receta creada correctamente')
            return redirect('dashboard')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': form.errors})
    else:
        form = RecetaForm()
    return render(request, 'gestion/form_modal.html', {'form': form, 'title': 'Crear Receta'})

@login_required
def receta_update(request, pk):
    receta = get_object_or_404(Receta, pk=pk)
    if request.method == 'POST':
        form = RecetaForm(request.POST, instance=receta)
        if form.is_valid():
            form.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Receta actualizada correctamente'})
            messages.success(request, 'Receta actualizada correctamente')
            return redirect('dashboard')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': form.errors})
    else:
        form = RecetaForm(instance=receta)
    return render(request, 'gestion/form_modal.html', {'form': form, 'title': 'Editar Receta'})

@login_required
def receta_delete(request, pk):
    receta = get_object_or_404(Receta, pk=pk)
    if request.method == 'POST':
        receta.delete()
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': 'Receta eliminada correctamente'})
        messages.success(request, 'Receta eliminada correctamente')
        return redirect('dashboard')
    return render(request, 'gestion/confirm_delete.html', {'object': receta, 'title': 'Confirmar eliminación de Receta'})

@login_required
def dashboard_reportes(request):
    reportes = ConsultaReporte.objects.all()
    return render(request, 'gestion/dashboard_reportes.html', {'reportes': reportes})

@login_required
def ejecutar_reporte(request):
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' and request.method == 'POST':
        try:
            reporte_id = request.POST.get('reporte_id')
            fecha_input = request.POST.get('fecha_input', None)
            
            if not reporte_id:
                return JsonResponse({
                    'success': False,
                    'error': 'ID de reporte no proporcionado'
                }, status=400)

            reporte = ConsultaReporte.objects.get(pk=reporte_id)
            sql = reporte.sql

            # Validar que la consulta sea SELECT
            if not sql.strip().lower().startswith('select'):
                return JsonResponse({
                    'success': False,
                    'error': 'Solo se permiten consultas SELECT'
                }, status=400)

            # Ejecutar la consulta
            with connection.cursor() as cursor:
                if '%s' in sql:
                    if not fecha_input:
                        return JsonResponse({
                            'success': False,
                            'error': 'Se requiere una fecha para este reporte'
                        }, status=400)
                    # Validar formato de fecha (YYYY-MM-DD)
                    if not re.match(r'^\d{4}-\d{2}-\d{2}$', fecha_input):
                        return JsonResponse({
                            'success': False,
                            'error': 'Formato de fecha inválido (use YYYY-MM-DD)'
                        }, status=400)
                    print(f"Ejecutando SQL: {sql} con fecha: {fecha_input}")  # Depuración
                    cursor.execute(sql, [fecha_input])  # Usar lista para %s
                else:
                    print(f"Ejecutando SQL: {sql}")  # Depuración
                    cursor.execute(sql)
                
                # Obtener columnas y filas
                columns = [col[0] for col in cursor.description]
                rows = cursor.fetchall()

            return JsonResponse({
                'success': True,
                'columns': columns,
                'rows': [list(row) for row in rows]
            })
        except ConsultaReporte.DoesNotExist:
            print(f"Reporte no encontrado: ID {reporte_id}")  # Depuración
            return JsonResponse({
                'success': False,
                'error': 'Reporte no encontrado'
            }, status=404)
        except Exception as e:
            print(f"Error en ejecutar_reporte: {str(e)}")  # Depuración
            return JsonResponse({
                'success': False,
                'error': f'Error al ejecutar el reporte: {str(e)}'
            }, status=500)
    print("Solicitud inválida: no es AJAX o no es POST")  # Depuración
    return JsonResponse({'success': False, 'error': 'Solicitud inválida'}, status=400)

@login_required
def exportar_reporte_pdf(request):
    if request.method == 'POST':
        try:
            reporte_id = request.POST.get('reporte_id')
            fecha_input = request.POST.get('fecha_input', None)
            
            if not reporte_id:
                return JsonResponse({
                    'success': False,
                    'error': 'ID de reporte no proporcionado'
                }, status=400)

            reporte = ConsultaReporte.objects.get(pk=reporte_id)
            sql = reporte.sql

            # Validar que la consulta sea SELECT
            if not sql.strip().lower().startswith('select'):
                return JsonResponse({
                    'success': False,
                    'error': 'Solo se permiten consultas SELECT'
                }, status=400)

            # Ejecutar la consulta
            with connection.cursor() as cursor:
                if '%s' in sql:
                    if not fecha_input:
                        return JsonResponse({
                            'success': False,
                            'error': 'Se requiere una fecha para este reporte'
                        }, status=400)
                    cursor.execute(sql, [fecha_input])
                else:
                    cursor.execute(sql)
                
                columns = [col[0] for col in cursor.description]
                rows = cursor.fetchall()

            # Crear PDF
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            elements = []
            
            # Estilos
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Centrado
            )
            
            # Título
            title = Paragraph(f"Reporte: {reporte.nombre}", title_style)
            elements.append(title)
            
            # Fecha de generación
            fecha_generacion = datetime.now().strftime("%d/%m/%Y %H:%M")
            fecha_para = Paragraph(f"Generado el: {fecha_generacion}", styles['Normal'])
            elements.append(fecha_para)
            elements.append(Spacer(1, 20))
            
            if fecha_input:
                fecha_filtro = Paragraph(f"Fecha filtro: {fecha_input}", styles['Normal'])
                elements.append(fecha_filtro)
                elements.append(Spacer(1, 20))

            # Crear tabla
            if rows:
                # Preparar datos para la tabla
                table_data = [columns]  # Headers
                for row in rows:
                    table_data.append([str(cell) if cell is not None else '-' for cell in row])
                
                # Crear tabla
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(table)
            else:
                no_data = Paragraph("No se encontraron datos para este reporte.", styles['Normal'])
                elements.append(no_data)

            # Construir PDF
            doc.build(elements)
            
            # Preparar respuesta
            buffer.seek(0)
            filename = f"reporte_{reporte.nombre.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            
            response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response

        except ConsultaReporte.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Reporte no encontrado'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error al generar PDF: {str(e)}'
            }, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

@login_required
def exportar_reporte_excel(request):
    if request.method == 'POST':
        try:
            reporte_id = request.POST.get('reporte_id')
            fecha_input = request.POST.get('fecha_input', None)
            
            if not reporte_id:
                return JsonResponse({
                    'success': False,
                    'error': 'ID de reporte no proporcionado'
                }, status=400)

            reporte = ConsultaReporte.objects.get(pk=reporte_id)
            sql = reporte.sql

            # Validar que la consulta sea SELECT
            if not sql.strip().lower().startswith('select'):
                return JsonResponse({
                    'success': False,
                    'error': 'Solo se permiten consultas SELECT'
                }, status=400)

            # Ejecutar la consulta
            with connection.cursor() as cursor:
                if '%s' in sql:
                    if not fecha_input:
                        return JsonResponse({
                            'success': False,
                            'error': 'Se requiere una fecha para este reporte'
                        }, status=400)
                    cursor.execute(sql, [fecha_input])
                else:
                    cursor.execute(sql)
                
                columns = [col[0] for col in cursor.description]
                rows = cursor.fetchall()

            # Crear Excel
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = reporte.nombre[:31]  # Límite de caracteres para nombre de hoja
            
            # Configurar estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Agregar información del reporte
            worksheet['A1'] = f"Reporte: {reporte.nombre}"
            worksheet['A1'].font = Font(bold=True, size=14)
            worksheet['A2'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            
            if fecha_input:
                worksheet['A3'] = f"Fecha filtro: {fecha_input}"
                start_row = 5
            else:
                start_row = 4
            
            # Agregar headers
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=start_row, column=col_num)
                cell.value = column_title
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Agregar datos
            if rows:
                for row_num, row_data in enumerate(rows, start_row + 1):
                    for col_num, cell_value in enumerate(row_data, 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = cell_value if cell_value is not None else '-'
                        cell.alignment = Alignment(horizontal="center")
            
            # Ajustar ancho de columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Preparar respuesta
            buffer = io.BytesIO()
            workbook.save(buffer)
            buffer.seek(0)
            
            filename = f"reporte_{reporte.nombre.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response

        except ConsultaReporte.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Reporte no encontrado'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error al generar Excel: {str(e)}'
            }, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)